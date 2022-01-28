from genericpath import exists
import pandas as pd
from datetime import date
from tkinter import HORIZONTAL, LEFT, S, Entry, OptionMenu, StringVar, Tk, Frame, Button, Label, messagebox, ttk
import os
import pendulum
from openpyxl import load_workbook, styles
import json

class TPR_Reporter:
    
    def __init__(self):
        # savedDay = self.getSavedDay()
        self.dayOfWeek = self.chooseDay()
        self.nextSaturday = self.getNextReportDate(pendulum.now())
        self.nextThreeSaturdays = self.getNextThreeSaturdays()
        self.nextSaturdayDateFormatted = self.formatDates(self.nextSaturday)
        self.brdFile = self.findBRDataFile()
        
    def chooseDay(self, day=''):
        todaysDate = pendulum.now()
        if day == 'Monday':
            return todaysDate.next(pendulum.MONDAY)
        elif day == 'Tuesday':
            return todaysDate.next(pendulum.TUESDAY)
        elif day == 'Wednesday':
            return todaysDate.next(pendulum.WEDNESDAY)
        elif day == 'Thursday':
            return todaysDate.next(pendulum.THURSDAY)
        elif day == 'Friday':
            return todaysDate.next(pendulum.FRIDAY)
        elif day == 'Saturday':
            return todaysDate.next(pendulum.SATURDAY)
        elif day == 'Sunday':
            return todaysDate.next(pendulum.SUNDAY)
        else:
            return todaysDate.next(pendulum.SATURDAY)
        
    def findBRDataFile(self):
        desktop = os.path.expanduser("~/Desktop")
        fileName = "BRdata_Prices.xlsx"
        file = os.path.join(desktop, fileName)
        self.checkUpdated(file)
        return file
    
    def formatDates(self, date):
        dateFormat = '%#m/%#d/%Y'
        return date.strftime(dateFormat)
    
    def getNextReportDate(self, date):
        return date.next(pendulum.SATURDAY)
    
    def checkUpdated(self, file):
        errorMsg = f"""
        {file}
         
        File does not exist or is outdated.
        
        Please run 'Parse BRData Prices' 
        before running this program.
        """
        
        if exists(file):
            fileModifiedDate = date.fromtimestamp(os.path.getmtime(file))
        else:
            messagebox.showerror("FileNotFound",
                                 errorMsg)
            exit()
            
        if fileModifiedDate == date.today():
            return
        else:
            messagebox.showerror("FileNotUpdated",
                                 errorMsg)
            exit()            
        
    def getData(self):
        tprColumn = 'TPR\nPrior'
        self.dateColumnName = 'TPR To'
        tprPriority = 99
        columnsToUse = "B,C,D,F,G,H,I,J,K,T"
        columnRenaming = {'Description': 'Item Description',
                               'Reg\nPM': ' ', 
                               'Reg\nPrice': 'Regular Price', 
                               'TPR\nPM': '  ',
                               'TPR\nPrice': 'TPR Price'}
        rawFile = pd.read_excel(self.brdFile, usecols=columnsToUse)
        filterOld = rawFile[rawFile[self.dateColumnName] >= self.nextSaturdayDateFormatted]
        filterFirstSat = filterOld[filterOld[self.dateColumnName] != self.nextThreeSaturdays[0]]
        filterSecondSat = filterFirstSat[filterFirstSat[self.dateColumnName] != self.nextThreeSaturdays[1]]
        filterThirdSat = filterSecondSat[filterSecondSat[self.dateColumnName] != self.nextThreeSaturdays[2]]
        tprFilter = filterThirdSat[filterThirdSat[tprColumn] == tprPriority]
        finalDataFile = tprFilter
        finalDataFile.rename(columns=columnRenaming,
                                        inplace=True)
        self.dataFile = finalDataFile

    def getNextThreeSaturdays(self):
        saturdays = []
        nextSat = self.nextSaturday
        for i in range(1,4):
            nextSat = self.getNextReportDate(nextSat)
            saturdays.append(self.formatDates(nextSat))
        return(saturdays)
                
    def setupFiles(self):
        fileDateFormat ='%#m-%#d-%Y'
        nxtSatDate = self.nextSaturday.strftime(fileDateFormat)
        self.reportFolder = os.path.expanduser("~\Desktop\TPR Report")
        if not os.path.exists(self.reportFolder):
            os.mkdir(self.reportFolder)
        self.reportFile = os.path.join(self.reportFolder,
                                       f'TPRreport{nxtSatDate}.xlsx')
        self.reportWriter = pd.ExcelWriter(self.reportFile,
                                           engine='xlsxwriter') 
              
    def createReport(self):
            self.getData()
            self.setupFiles()
            self.createSheets()
            self.reportWriter.close()
            self.postProcessing()
            self.completedReports()  
    
    def completedReports(self):
        infoText = """
            Report created and stored in the
            TPR Report folder located on your desktop.
                    """
        messagebox.showinfo(title="Report Compiled",
                            message=infoText)

    def createSheets(self):
        upcString = '[<=99999]#;[<=9999999999]#####-#####;###-#####-#####'
        workbookFormats = {'upc':{'num_format': upcString},
                           'num':{'num_format': '$#.00'}}
        departments = {'Produce':[20,21,22,23,24,25,26,27,28,29,100],
                            'Meat':[30,31,32,33,34,35,36,37,38,39],
                            'Frozen':[40,41,42,43,44,45,46,47,48,49],
                            'Dairy':[50,51,52,53,54,55,56,57,58,59],
                            'Deli & Bakery':[60,61,62,63,64,65,66,67,68,69,
                                             80,81,82,83,84,85,86,87,88,89],
                            'GM & HBC':[70,71,72,73,74,75,76,77,78,79,
                                        90,91,92,93,94,95,96,97,98,99],
                            'Grocery':[200,201,202,203,204,205,
                                       206,207,208,209,210],
                            'Stray TPRs':range(1,211)}
        columnsList = ["UPC",
                       "Item Description",
                       " ",
                       "Regular Price",
                       "  ",
                       "TPR Price"]
        for dept in departments.keys():
            if dept != 'Stray TPRs':
                self.processDepts(dept, departments,columnsList, workbookFormats)
            else:
                self.processDepts(dept, departments,columnsList, workbookFormats, strays=True)
                
    def processDepts(self, dept, departments,columnsList, workbookFormats, strays=False):
        regularTPRs = self.dataFile[self.dataFile[self.dateColumnName] == self.nextSaturdayDateFormatted]
        strayTPRS = self.dataFile[self.dataFile[self.dateColumnName] != self.nextSaturdayDateFormatted]
        numList = departments[dept]
        if not strays:
            departmentTPRs = regularTPRs[regularTPRs['Dept'].isin(numList)]
        else:
            departmentTPRs = strayTPRS[strayTPRS['Dept'].isin(numList)]
        if not departmentTPRs.empty:
            sortedTPRs = departmentTPRs.sort_values(by=['UPC'])
            rows = sortedTPRs.index
            sortedTPRs.to_excel(self.reportWriter, 
                        sheet_name=dept,
                        index=False,
                        columns=columnsList)
            reportWorkbook = self.reportWriter.book
            moneyFormat = reportWorkbook.add_format(workbookFormats['num'])
            upcFormat = reportWorkbook.add_format(workbookFormats['upc'])
            reportWorksheet = self.reportWriter.sheets[dept]
            if dept != 'Stray TPRs':
                headerFormat = (f'&C&20TPR Report  |'
                                f'|  {dept} Department  |'
                                f'|  {self.nextSaturdayDateFormatted}')
            else:
                headerFormat = (f'&C&20TPR Report  |'
                                f'|  {dept} All Departments  |'
                                f'|  {self.nextSaturdayDateFormatted}')
            reportWorksheet.set_header(headerFormat)
            moneyFormat.set_align('center')
            reportWorksheet.set_column('A:A', 14.86, upcFormat)
            reportWorksheet.set_column('B:B', 38)
            reportWorksheet.set_column('C:C', 2.29)
            reportWorksheet.set_column('D:D', 11.86, moneyFormat)
            reportWorksheet.set_column('E:E', 2.29)
            reportWorksheet.set_column('F:F', 8.43, moneyFormat)
            
    def addBorders(self):
        self.processSheets()
           
    def processSheets(self):
        workbook = load_workbook(self.reportFile)
        for sheets in workbook.sheetnames:
            sheet = workbook[sheets]
            self.setBorder(sheet)
        workbook.save(self.reportFile )
        workbook.close()
      
    def setBorder(self, worksheet):
        workSheetRange = ['A', 'F']
        thinFormat = styles.Side(border_style="dotted", color="000000")
        addBorder = styles.Border(bottom=thinFormat)
        currentSheet = worksheet[workSheetRange[0]:workSheetRange[1]]
        for row in currentSheet:
            for cell in row: 
                if cell.row % 2 != 0 and cell.row != 1:
                    cell.border = addBorder

    def postProcessing(self):
        self.addBorders()
     
        
class TPR_Reporter_GUI:
    
    # Initialization
    def __init__(self):
        self.settings = dict()
        self.settingsInit()
        mainWindow = self.setupDisplay()
        mainWindow.mainloop()
    
    # Init Variables
    def comboBoxVariables(self):
        self.dayOfWeek  = StringVar()
        self.dayOfMonth = StringVar()
        self.reportFrequency = StringVar()

        
    # Create the main window.   
    def createWindow(self):
        window = Tk()
        window.title("TPR Report")
        window.geometry("400x300")
        return window
    
    # Creates a master frame for the window.
    def createFrame(self, window):
        mainFrame = Frame(window)
        return mainFrame

    # Init's the main frame/window, calls the widgets, and returns the mainwindow.
    def setupDisplay(self):
        mainWindow = self.createWindow()
        mainFrame = Frame(mainWindow)
        self.comboBoxVariables()
        self.setupWidgets(mainFrame)
        mainFrame.pack()
        return mainWindow
    
    # Calls each widget to be set up.
    def setupWidgets(self, frame):
        self.setupFrequencyControls(frame)
        # frequencyFrame = self.frequencyFrame(frame)
        self.titleLabel(frame)
        # self.chooseFrequency(frequencyFrame)
        # self.nextReportLabel(frame)
        self.compileButtonMethod(frame)
        self.pleaseWaitLabel(frame)
        self.finishedLabelMethod(frame)
        # labsTest = Label(frequencyFrame, text='test')
        # labsTest.pack()
        # self.frequencyDecider(frequencyFrame)
        # self.dayOfWeekMenu(frequencyFrame)
        
    # Frequency Controls
    # Freq 1/
    def setupFrequencyControls(self,frame):
        freqFrame = self.frequencyFrame(frame)
        self.dayOfWeekMenu(freqFrame)
        self.dayOfMonthControls(freqFrame)
        self.frequencyDecider(freqFrame)
        
    # Main Freq frame
    # Freq 2/        
    def frequencyFrame(self, frame):
        frequencyLabelFrame = ttk.Labelframe(frame,
                                             text="Options",
                                             padding=10)
        frequencyLabelFrame.pack(fill='both', expand='yes', side=LEFT)
        return frequencyLabelFrame
    
    # Combo box to select frequency of reports.
    # Freq 3/
    def frequencyDecider(self,frame):
        freqLabelText = "TPR Report Frequency"
        freqLabel = Label(frame, text=freqLabelText)
        freqLabel.pack()
        options = ["Weekly", "Monthly"]
        self.reportFrequency.set(self.settings["Frequency"])
        comboFrequency = ttk.Combobox(frame, 
                                      textvariable=self.reportFrequency,
                                values=options,
                                state="readonly",
                                width=10)
        comboFrequency.bind("<<ComboboxSelected>>", self.updateFrequencySelection)
        comboFrequency.pack()
        freqSep = ttk.Separator(frame, orient=HORIZONTAL)
        freqSep.pack()
        self.updateFrequencySelection("pass")
    
    def updateFrequencySelection(self,event):
        option = self.reportFrequency.get()
        
        if option == "Weekly":
            self.dayEntryFrame.pack_forget()
            self.dropdownMenu.pack()
        elif option == "Monthly":
            self.dropdownMenu.pack_forget()
            self.dayEntryFrame.pack()  
    # def chooseFrequency(self, frame):
    #     sep = ttk.Separator(frame, orient=HORIZONTAL)
    #     sep.pack(fill='x')
    #     lblFrame = ttk.Labelframe(frame, text="label")
    #     lblFrame.pack(fill='both', expand='yes', side=LEFT)
    #     ladfsd = Label(lblFrame, text='Test')   
    #     ladfsd.pack() 
    #     # comboFrequency = ttk.Combobox(frame,
    #     #                               values=[
    #     #                                         "Weekly",
    #     #                                         "Monthly"],
    #     #                               state="readonly")
    #     # comboFrequency.pack()
    
    # If Weekly selected then display these options
    # Freq 4/  
    def dayOfWeekMenu(self, frame):
        options = ["Monday", "Tuesday", "Wednesday",
                   "Thursday", "Friday", "Saturday", "Sunday"]
        
        self.dayOfWeek.set(self.settings["DayOfWeek"])
        self.dropdownMenu = ttk.Combobox(frame, 
                                    textvariable=self.dayOfWeek, 
                                    values=options,
                                    state="readonly",
                                    width=12)
  
    def dayOfMonthControls(self,frame):
        self.dayEntryFrame = Frame(frame)
        dayLabelText = "Day of Month: "
        self.dayLabel = Label(self.dayEntryFrame, text=dayLabelText)
        self.dayEntry = Entry(self.dayEntryFrame, width=2, 
                              textvariable=self.dayOfMonth)
        self.dayLabel.pack(side=LEFT)
        self.dayEntry.pack(side=LEFT)
        
    
    def nextReportLabel(self, frame):
        nextReportDay = self.getNextReportDay()
        labelText = f"Next {self.dayOfWeek.get()} is: {nextReportDay}"
        labelReportDay = Label(frame, text=labelText)
        
        labelReportDay.pack()
    
    def getNextReportDay(self):
        return TPR_Reporter().nextSaturdayDateFormatted
        
    def titleLabel(self, frame):
        titleText = "Press 'Compile Report' to get started."
        labelText = Label(frame, text=titleText)
        
        labelText.pack()
        
    def compileButtonMethod(self, frame):
        buttonText = 'Compile Report'
        self.compileButton = Button(frame, text=buttonText)
        
        self.compileButton.bind("<Button-1>", self.compileReports)
        self.compileButton.pack(anchor = S)
        
    def pleaseWaitLabel(self, frame):
        labeltext = "Processing.  Please be patient.  This may take a minute."
        self.waitLabel = Label(frame, text=labeltext)
    
    def finishedLabelMethod(self, frame):
        labelText = "Report compiled.\n  Please check the TPR report folder on your desktop"
        self.finishedLabel = Label(frame, text=labelText)
          
    def compileReports(self, event):
        self.compileButton.pack_forget()
        self.waitLabel.pack()
        TPR_Reporter().createReport()
        self.waitLabel.pack_forget()
        self.finishedLabel.pack()
        

         
    def settingsInit(self):
        self.settingsFile = 'TPR_Report_Config.json'
        if exists(self.settingsFile):
            pass
        else:
            with open(self.settingsFile, "w") as output:
                settings = dict()
                settings['DayOfWeek'] = 'Saturday'
                settings['Frequency'] = 'Weekly'
                settingsJson = json.dumps(settings)
                output.write(settingsJson)
        self.loadSettings()
                
    def saveSettings(self, event):
        settingsJson = json.dumps(self.settings)       
        with open(self.settingsFile, "w") as output:
            output.write(settingsJson)
            
    def loadSettings(self):
        with open(self.settingsFile) as input:
            settings = json.load(input)
            for item in settings:
                self.settings[item] = settings[item]
                

if __name__ == "__main__":
    instance = TPR_Reporter_GUI()
    